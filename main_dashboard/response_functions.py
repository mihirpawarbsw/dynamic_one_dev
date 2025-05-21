import json
import os
import re
import time
import ast
import pandas as pd
import numpy as np
# from django.shortcuts import render
# from django.http import HttpResponse, JsonResponse
# from django.shortcuts import render, redirect
# from django.conf import settings
# # Create your views here.
# # from CCV_Tool.CCV_Tool.settings import pythonpath
from django.conf import settings

from datetime import date
from datetime import datetime
from io import StringIO

# import win32com.client as win32
# from PIL import Image
# from fpdf import FPDF
from openpyxl.styles import numbers, PatternFill, Border, Side, Font,NamedStyle
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from openpyxl import Workbook

# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import letter
from openpyxl.styles import Alignment, Font

# output_pythonpath = r"C:\Mihir_Python_Projects\BCST_Tool_Core_Python_Logic\ccv_tool_Sales\output_crosstab\\"
# PYTHONPATH = r"C:\Users\MihirPawar\Desktop\Python Project BSW\Python Files1\ccv_tool\ccv_survey_data\electrolux\\"
# PYTHONPATH = r"C:\Users\MihirPawar\Desktop\Python Project BSW\Python Files1\ccv_tool\ccv_survey_data\\"
# MERGED_PYTHONPATH = merged_pythonpath = r"C:\Users\MihirPawar\Desktop\Python Project BSW\Python Files1\ccv_tool\merged_data_files\\"

def allign_grand_total_headers_fn_dynamic(cross_df):
    Columns_df_lst = []
    
    # Determine the number of levels in the multi-index columns
    num_levels = cross_df.columns.nlevels
    
    for i in range(num_levels):
        level_names = cross_df.columns.get_level_values(i).unique().tolist()
        cross_df_grp = cross_df.groupby(axis=1, level=i)
        
        for name in level_names:
            cross_df_col = cross_df_grp.get_group(name)
            cross_df_col_df = pd.DataFrame(cross_df_col)
            next_level = i + 1
            if next_level < num_levels:
                unique_groups_next_level = cross_df_col_df.columns.get_level_values(next_level).unique().tolist()
                
                if 'Grand Total' in unique_groups_next_level:
                    unique_groups_next_level.remove('Grand Total')
                    unique_groups_next_level.insert(0, 'Grand Total')
                    
                cross_df_col_df_reordered = cross_df_col_df.loc[:, cross_df_col_df.columns.get_level_values(next_level).isin(unique_groups_next_level)]
                
                # Remove duplicate columns
                cross_df_col_df_reordered = cross_df_col_df_reordered.loc[:, ~cross_df_col_df_reordered.columns.duplicated()]
                
                cross_df_col_df_reordered = cross_df_col_df_reordered.reindex(columns=unique_groups_next_level, level=next_level)
                
                Columns_df_lst.append(cross_df_col_df_reordered)
    
    if len(Columns_df_lst) > 0:
        cross_df = pd.concat(Columns_df_lst, axis=1)
    
        if 'Grand Total' in cross_df.columns.get_level_values(0).unique():
            grand_total_name = 'Grand Total'
            level_names_cols = cross_df.columns.get_level_values(0).unique().tolist()
            level_names_cols.remove(grand_total_name)
            level_names_cols.insert(0, grand_total_name)
            cross_df = cross_df.reindex(columns=level_names_cols, level=0)
    else:
        cross_df = pd.DataFrame()
    
    return cross_df

############ ADDED ON 04-10-2024 - ALLIGNING TOTALS #####################################
############ ADDED ON 04-10-2024 - ALLIGNING TOTALS #####################################
def allign_grand_total_headers_fn(cross_df):

    Columns_df_lst = []
    # Get the level 0 names in the columns
    level_0_names_cols = cross_df.columns.get_level_values(0).unique().tolist()

    cross_df_col_grp = cross_df.groupby(axis=1, level=0)

    # Access the first group
    for cross_df_col_loop in level_0_names_cols:
        cross_df_col = cross_df_col_grp.get_group(cross_df_col_loop)
        cross_df_col_df = pd.DataFrame(cross_df_col)

        unique_groups_level1 = cross_df_col_df.columns.get_level_values(1).unique().tolist()
        level0_grp_names = cross_df_col_df.columns.get_level_values(0).unique().tolist()
        # print('level0_grp_names', level0_grp_names)
        # print('unique_groups_level1', unique_groups_level1)

        if 'Total (Among Displayed)' in unique_groups_level1:
            unique_groups_level1.remove('Total (Among Displayed)')
            unique_groups_level1.append('Total (Among Displayed)')
        else:
            pass
            
        if 'Grand Total' in unique_groups_level1:
            # Remove 'Total' from the original list
            unique_groups_level1.remove('Grand Total')
            

            # Insert 'Total' at the first position
            # unique_groups_level1.insert(-1, 'Grand Total')
            
            unique_groups_level1.append('Grand Total')
            # print('unique_groups_level1=====',unique_groups_level1)
        else:
            pass

        cross_df_col_df_reordered = cross_df_col_df.loc[:,
                      cross_df_col_df.columns.get_level_values(1).isin(list(unique_groups_level1))]
        # cross_df_col_df_reordered.to_excel('cross_df_col_df_reordered_ggg.xlsx')

        try:
            # print('====688 duplicate')
            cross_df_col_df_reordered = cross_df_col_df_reordered.loc[:, ~cross_df_col_df_reordered.columns.duplicated(keep='first')]
        except:
            pass
        ########################################################################################
        # cross_df_col_df_reordered = cross_df_col_df_reordered.iloc[:, [0, 1]]
        cross_df = cross_df_col_df_reordered.reindex(columns=unique_groups_level1, level=1)
        # cross_df.to_excel('cross_df_col_df_reordered_ffff.xlsx')

        Columns_df_lst.append(cross_df)
    cross_df = pd.concat(Columns_df_lst,axis=1)

    ########################## ADDED ON 23-01-2024 #####################################
    if 'Grand Total' in level_0_names_cols:

        # Move "Total" to the first position in the list
        level_0_names_cols.remove("Grand Total")
        level_0_names_cols =  level_0_names_cols + ["Grand Total"]

        cross_df = cross_df.reindex(columns=level_0_names_cols, level=0)

    else:
        cross_df = cross_df.copy()

    ########################## ADDED ON 23-01-2024 #####################################
    # cross_df.to_excel('CROSS_DF_Before_return.xlsx')
    return cross_df
############ ADDED ON 04-10-2024 - ALLIGNING TOTALS #####################################
############ ADDED ON 04-10-2024 - ALLIGNING TOTALS #####################################

############################ added on 24-09-2024 ###########################################

def allign_grand_total_headers_fn_OLD_04102024_BK(cross_df):
    start_time = time.time()
    # Initialize list to store DataFrames for each group
    Columns_df_lst = []

    # Get unique level 0 names in the columns
    level_0_names_cols = cross_df.columns.get_level_values(0).unique().tolist()

    # Group columns by level 0
    cross_df_col_grp = cross_df.groupby(axis=1, level=0)

    # Iterate over unique level 0 names
    for cross_df_col_loop in level_0_names_cols:
        cross_df_col = cross_df_col_grp.get_group(cross_df_col_loop)
        unique_groups_level1 = cross_df_col.columns.get_level_values(1).unique().tolist()

        # Check and adjust for 'Grand Total'
        if 'Grand Total' in unique_groups_level1:
            unique_groups_level1.remove('Grand Total')
            unique_groups_level1.append('Grand Total')

        # Reorder columns based on unique level 1 names
        cross_df_col_df_reordered = cross_df_col.loc[:, cross_df_col.columns.get_level_values(1).isin(unique_groups_level1)]
        
        # Drop duplicate columns, keeping the first
        cross_df_col_df_reordered = cross_df_col_df_reordered.loc[:, ~cross_df_col_df_reordered.columns.duplicated(keep='first')]

        # Reindex to ensure 'Grand Total' is at the end
        cross_df_col_df_reordered = cross_df_col_df_reordered.reindex(columns=unique_groups_level1, level=1)
        
        Columns_df_lst.append(cross_df_col_df_reordered)

    # Concatenate all processed DataFrames
    cross_df = pd.concat(Columns_df_lst, axis=1)

    # Handle the 'Grand Total' at the level 0
    if 'Grand Total' in level_0_names_cols:
        level_0_names_cols.remove("Grand Total")
        level_0_names_cols.append("Grand Total")
        cross_df = cross_df.reindex(columns=level_0_names_cols, level=0)

    end_time = time.time()
    print('Time taken to REORDER GRAND TOTAL -',end_time-start_time,' seconds')
    return cross_df

############################ added on 24-09-2024 ###########################################
def allign_grand_total_headers_fn_old(cross_df):

    Columns_df_lst = []
    # Get the level 0 names in the columns
    level_0_names_cols = cross_df.columns.get_level_values(0).unique().tolist()

    cross_df_col_grp = cross_df.groupby(axis=1, level=0)

    # Access the first group
    for cross_df_col_loop in level_0_names_cols:
        cross_df_col = cross_df_col_grp.get_group(cross_df_col_loop)
        cross_df_col_df = pd.DataFrame(cross_df_col)

        unique_groups_level1 = cross_df_col_df.columns.get_level_values(1).unique().tolist()
        level0_grp_names = cross_df_col_df.columns.get_level_values(0).unique().tolist()
        # print('level0_grp_names', level0_grp_names)
        # print('unique_groups_level1', unique_groups_level1)

        if 'Grand Total' in unique_groups_level1:
            # Remove 'Total' from the original list
            unique_groups_level1.remove('Grand Total')

            # Insert 'Total' at the first position
            # unique_groups_level1.insert(-1, 'Grand Total')
            unique_groups_level1.append('Grand Total')
            # print('unique_groups_level1=====',unique_groups_level1)
        else:
            pass

        cross_df_col_df_reordered = cross_df_col_df.loc[:,
                      cross_df_col_df.columns.get_level_values(1).isin(list(unique_groups_level1))]
        # cross_df_col_df_reordered.to_excel('cross_df_col_df_reordered_ggg.xlsx')

        try:
            # print('====688 duplicate')
            cross_df_col_df_reordered = cross_df_col_df_reordered.loc[:, ~cross_df_col_df_reordered.columns.duplicated(keep='first')]
        except:
            pass
        ########################################################################################
        # cross_df_col_df_reordered = cross_df_col_df_reordered.iloc[:, [0, 1]]
        cross_df = cross_df_col_df_reordered.reindex(columns=unique_groups_level1, level=1)
        # cross_df.to_excel('cross_df_col_df_reordered_ffff.xlsx')

        Columns_df_lst.append(cross_df)
    cross_df = pd.concat(Columns_df_lst,axis=1)

    ########################## ADDED ON 23-01-2024 #####################################
    if 'Grand Total' in level_0_names_cols:

        # Move "Total" to the first position in the list
        level_0_names_cols.remove("Grand Total")
        level_0_names_cols =  level_0_names_cols + ["Grand Total"]

        cross_df = cross_df.reindex(columns=level_0_names_cols, level=0)

    else:
        cross_df = cross_df.copy()

    ########################## ADDED ON 23-01-2024 #####################################
    # cross_df.to_excel('CROSS_DF_Before_return.xlsx')
    return cross_df
########################## ADDED ON 30-01-2024 ############################################################

def create_selected_weight_columns_old(selected_measure_columns,df):

    # print('dfrrrr',df)
    # selected_measure_columns = ['Sales (LC)']
    df = df[df['Variable'].isin(selected_measure_columns)]
    df = df[df['Count'] > 1]
    # print('SHISHIEDO_SELL_OUT',df)

    # Extract values from the 'Codeframe' column and convert to a single list
    selected_weight_column_all = []
    for value in df['Codeframe']:
        # Use ast.literal_eval to safely evaluate the literal expression
        codeframe_list = ast.literal_eval(value)
        selected_weight_column_all.extend(codeframe_list)

    # print('selected_weight_column_all',selected_weight_column_all)
    
    return selected_weight_column_all

####################### BASE FILTER #####################################################################

#################################### added on 18-02-2025 #####################################
def base_filter_data(df, dict_base_filter_data):
    for key, value in dict_base_filter_data.items():
        df = df[df[key].isin(value)]
    return df
#################################### added on 18-02-2025 #####################################


def base_filter_resp(df, dict_base_filter, row_name, col_name):
    # print('======== base_filter_resp =======')
    # print('row_name===', row_name)
    # print('col_name===', col_name)
    ####################### BASE FILTER #####################################################################
    # dict_base_filter = {

    #     'Electrolux_India_respondent':['Gender','Age_:_Post_code'],
    #     # 'Electrolux_China_respondent': ['Gender']
    # }

    keys = list(dict_base_filter.keys())[0]
    values = list(dict_base_filter.values())[0]
    # print('keys 211232',keys)
    # print('values 211',values)
    # print('values 211 TYPE',type(values))

    if 'Time' in values:
        values.remove('Time')
    #################### OTHER VARIABLE FILTER LOGIC COMMENTED - 08-04-2024 #############
    # print('values beforee==', values)
    # print('values beforee==', values)
    # # sum_row_col = row_name + col_name + selected_weight_column_all
    # sum_row_col = row_name + col_name
    # print('sum_row_col===', sum_row_col)
    # values = [[x for x in values if x not in sum_row_col]]
    # print('values after===', values)
    #################### OTHER VARIABLE FILTER LOGIC COMMENTED - 08-04-2024 #############
    #
    filter_dict_resp = {}

    # table_name = table_name + ".json"
    table_name = keys + ".xlsx"

    # df = pd.read_json(settings.PYTHONPATH + table_name, orient='records', lines=True)
    df = df[values]

    for selected_column in df.columns:
        categories_list = df[selected_column].unique().tolist()
        # categories_list = sorted(df[selected_column].unique())
        # print('Column name-')
        # print(selected_column,'Count-',len(categories_list))
        # df_Freq = pd.DataFrame(columns=['Count'])

        # freq_vals = df[selected_column].value_counts()

        # df_Freq['Count'] = freq_vals
        # df_Freq = df_Freq.reset_index().rename(columns={'index': 'Column_data'})
        # df_Freq['Column_data'] = df_Freq['Column_data'].astype(str)
        # df_Freq = df_Freq.sort_values('Column_data')
        # print("==df_Freq==", df_Freq)

        # categories_list = df_Freq['Column_data']

        filter_dict_resp11 = {selected_column: list(categories_list)}
        filter_dict_resp.update(filter_dict_resp11)
    # print('filter_dict_resp 250',filter_dict_resp)
    ####################### BASE FILTER #####################################################################

    # print('filter_dict_resp====', filter_dict_resp)
    # print('======== base_filter_resp =======')

    if 'Time' in list(filter_dict_resp.keys()):
        filter_dict_resp.pop('Time')

    return filter_dict_resp


#################### ADDED ON 25-07-2023 ##########################################

#################### ADDED ON 18-02-2025 ###############################################
def base_filter_resp_all(df, dict_table, row_name, col_name):
    filter_dict_resp = {}
 
    # Extract table names and filter conditions
    for table_name, values_temp in dict_table.items():
        # Filter the dataframe based on the values in dict_table
        df_filtered = df[values_temp]
 
        for selected_column in df_filtered.columns:
            # Get unique categories and prepend 'Total'
            categories_list = ['Total'] + df_filtered[selected_column].unique().tolist()
 
            # Update the dictionary with the column categories
            filter_dict_resp[selected_column] = categories_list
 
    return filter_dict_resp
#################### ADDED ON 18-02-2025 ###############################################

#################### ADDED ON 25-07-2023 ##########################################

def saving_crosstab_excel_file(cross_df, legends_df, percent_calc):
    print('FUNCTION saving_crosstab_excel_file BEGINS...')
    ####################################################### NEW CODE! ##################################
    cross_df_highlighter = cross_df.copy()
    cross_df_highlighter.to_excel(settings.FINAL_CROSSTAB_OUTPUT + 'cross_df_highlighter.xlsx', engine="openpyxl")
    legends_df.to_excel(settings.FINAL_CROSSTAB_OUTPUT + 'legends_df.xlsx', engine="openpyxl")

    workbook_top = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'cross_df_highlighter.xlsx')
    sheet_top = workbook_top.active

    # Open the bottom Excel file
    workbook_bottom = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'legends_df.xlsx')
    sheet_bottom = workbook_bottom.active

    # # Apply formatting to the top Excel file based on the calculation type
    # if percent_calc == 'column_percent' or percent_calc == 'row_percent' or percent_calc == 'table_percent' or percent_calc == 'Significance':
    #     percentage_format = numbers.FORMAT_PERCENTAGE_00
    #     for row in sheet_top.iter_rows():
    #         for cell in row:
    #             cell.number_format = percentage_format

    # Define a custom style with the number format
    # indian_number_style = NamedStyle(name="indian_number_style")
    # indian_number_style.number_format = '##,##,##0;##0'

    # # Loop through the cells containing numbers
    # for row in sheet_top.iter_rows():
    #     for cell in row:
    #         if isinstance(cell.value, (int, float)) and cell.value >= 1000:
    #             cell.style = indian_number_style

    ############# TABLE BORDER ###########################
    for row in sheet_top.iter_rows():
        for cell in row:
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            # Apply the border to the cell
            cell.border = border
    ############## TABLE BORDER ##########################

    ################# FOR INDICES ONLY ###########################################################

    # Get the maximum row index of the top Excel file
    max_row_top = sheet_top.max_row

    # Append the rows from the bottom Excel file to the top Excel file
    for row in sheet_bottom.iter_rows(min_row=2):
        new_row = [cell.value for cell in row]
        sheet_top.append(new_row)

    # Save the combined Excel file
    workbook_top.save(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_CROSSTAB_DF.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################
    workbook = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_CROSSTAB_DF.xlsx')
    # Select the active sheet
    sheet = workbook.active

    # Adjust column width TO 15
    for column in sheet.columns:
        # max_length = 0
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = 15

    # Adjust row height based on cell contents
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                lines = str(cell.value).count('\n') + 1
                cell_height = (lines * 14) + 6
                if cell_height > max_height:
                    max_height = cell_height
            except TypeError:
                pass
        sheet.row_dimensions[row[0].row].height = max_height

    # right align all cell values and remove bold formatting from text
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='right', vertical='top')
            cell.font = Font(bold=False)

    # Iterate through all cells in the current sheet
    for row in sheet.iter_rows():
        for cell in row:
            # Check if the cell contains "000AAAATotal"
            if isinstance(cell.value, str) and "000AAAATotal" in cell.value:
                # Replace "000AAAATotal" with "Grand Total"
                cell.value = cell.value.replace("000AAAATotal", "Grand Total")


    # Save the modified Excel file
    workbook.save(settings.FINAL_CROSSTAB_OUTPUT + 'cross_df.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################


################################### SAVING SIGNIFICANCE EXCEL FILE ##################################
def saving_siginficance_excel_file(significance_df):
    print('FUNCTION saving_siginficance_excel_file BEGINS...')
    ####################################################### NEW CODE! ##################################
    significance_df_highlighter = significance_df.copy()
    significance_df_highlighter.to_excel(settings.FINAL_CROSSTAB_OUTPUT + 'significance_df_highlighter.xlsx',
                                         engine="openpyxl")

    workbook_top = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    sheet_top = workbook_top.active

    ############# TABLE BORDER ###########################
    for row in sheet_top.iter_rows():
        for cell in row:
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            # Apply the border to the cell
            cell.border = border
    ############## TABLE BORDER ##########################

    # Get the maximum row index of the top Excel file
    max_row_top = sheet_top.max_row

    # Save the combined Excel file
    workbook_top.save(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################
    workbook = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    # Select the active sheet
    sheet = workbook.active

    # Adjust column width TO 15
    for column in sheet.columns:
        # max_length = 0
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = 15

    # Adjust row height based on cell contents
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                lines = str(cell.value).count('\n') + 1
                cell_height = (lines * 14) + 6
                if cell_height > max_height:
                    max_height = cell_height
            except TypeError:
                pass
        sheet.row_dimensions[row[0].row].height = max_height

    # Left align all cell values and remove bold formatting from text
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top')
            cell.font = Font(bold=False)

    # Save the modified Excel file
    workbook.save(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################

    ################################### SAVING SIGNIFICANCE EXCEL FILE ##################################


def codeframe(list_resp):
    # df = pd.read_json(PYTHONPATH + "Electrolux_India.json",orient='records', lines=True)
    # codeframe_df = pd.read_json(PYTHONPATH + "Electrolux- India_codeframe.json",orient='records', lines=True)
    df = pd.DataFrame(list_resp[0])
    codeframe_df = pd.DataFrame(list_resp[1])

    codeframe_df['Variable Level'].fillna('None', inplace=True)
    codeframe_df['Variable Name'].fillna('None', inplace=True)

    encode_dict = {}
    for loop_vals in range(len(codeframe_df)):
        codeframe_dict = {codeframe_df['Variable Name'][loop_vals]: codeframe_df['Variable Level'][loop_vals]}
        encode_dict.update(codeframe_dict)

    df.rename(columns=encode_dict, inplace=True)

    new_col_names_lst = []
    for loop in df.columns:
        # print("loop",loop)
        loop = loop.replace(" ", "_")
        new_col_names_lst.append(loop)

    df.columns = new_col_names_lst
    return df


def column_dtypes(list_all_files):
    colname_dict_final = {}
    for loop_filename in list_all_files:
        # pythonpath=r"C:/Users/MihirPawar/Desktop/Python Project BSW/Python Files1/ccv_tool//"
        df = pd.read_json(settings.PYTHONPATH + loop_filename, orient='records', lines=True)
        # print("df", df.head(2))

        # colname_obj = []
        # for x in df.columns:
        #         colname_obj.append(x)

        # colname_obj = [x for x in df.columns]
        # colname_obj = [x for x in df.columns.values]
        # colname_obj = list(df.columns.values)

        # print("colname_num", colname_num)
        # print("colname_obj", colname_obj)

        # Select numerical columns
        colname_num = df.select_dtypes(include=['int64', 'float64']).columns.tolist()

        # Select categorical columns
        colname_obj = df.select_dtypes(include=['object']).columns.tolist()

        col_dtype_dict = {'filename': loop_filename, 'colname_obj': colname_obj,
        'colname_num': colname_num}
        # filename11 = loop_filename.replace(".json", "")
        filename11 = loop_filename.replace(".xlsx", "")

        colname_dict_final1 = {filename11: col_dtype_dict}

        # print("col_dtype_dict",col_dtype_dict)
        colname_dict_final.update(colname_dict_final1)

    return colname_dict_final


def column_data_display(list_all_files):
    colname_data_dict_final = {}
    for loop_filename in list_all_files:
        df = pd.read_json(settings.PYTHONPATH + loop_filename, orient='records', lines=True)
        # print("df", df.head(2))

        colname_data_dict_final_22 = {}
        # selected_column='Age_Group'
        for selected_column in df.columns:
            # print("selected_column selected_column",selected_column)
            freq_percent = df[selected_column].value_counts(normalize=True).mul(100)
            ##print("freq_percent",freq_percent)

            freq_vals = df[selected_column].value_counts()

            df_Freq = pd.DataFrame(columns=['Count', 'Percent'])
            # df_Freq['Column']=df.columns
            df_Freq['Count'] = freq_vals
            df_Freq['Percent'] = freq_percent

            df_Freq['Count'] = round(df_Freq['Count'], 2)
            df_Freq['Percent'] = round(df_Freq['Percent'], 2)

            df_Freq['Percent'] = df_Freq['Percent'].astype(str) + " %"

            df_Freq = df_Freq.reset_index().rename(columns={'index': 'Column_data'})
            # print('df_Freq-->')
            # print('typee df_Freq',type(df_Freq))

            df_Freq_json1 = df_Freq.to_json(orient='index')
            df_Freq_json = json.loads(df_Freq_json1)

            # df_Freq.to_excel('df_Freq.xlsx')

            colname_data_dict_final_selected_column = {selected_column: df_Freq_json}
            colname_data_dict_final_22.update(colname_data_dict_final_selected_column)

        # filename22 = loop_filename.replace(".json", "")
        filename22 = loop_filename.replace(".xlsx", "")

        colname_data_dict_final1 = {filename22: colname_data_dict_final_22}
        # print(" ===== resp out of loop ===")
        # print("colname_data_dict_final1", colname_data_dict_final1)

        colname_data_dict_final.update(colname_data_dict_final1)
    # print("======= resp enddd =====")
    return colname_data_dict_final


def value_count_freq(df, filename, row_name, col_name):
    df = pd.read_json(settings.PYTHONPATH + filename, orient='records', lines=True)
    ##print("df", df.head(2))
    ##print("df colsss", df.columns)

    rowcount_dict = {}
    for selected_column in row_name:

        freq_vals = df[selected_column].value_counts()

        df_Freq = pd.DataFrame(columns=['Values'])
        # df_Freq['Column']=df.columns
        df_Freq['Values'] = freq_vals
        df_Freq = df_Freq.reset_index().rename(columns={'index': 'Columns'})

        # print('df_Freq colnames', df_Freq.columns)

        col_categories = df_Freq['Columns']

        # print('col_categories col_categories===', col_categories)

        df_Freq_json = dict(col_categories)

        # print('df_Freq shapeee', len(df_Freq))

        if selected_column == row_name[0]:
            temp_dict = {len(df_Freq): 'All'}

            df_Freq_json.update(temp_dict)

        # print("df_Freq_json df_Freq_json updatedd roww", df_Freq_json)

        rowcount_dict1 = {selected_column: df_Freq_json}
        rowcount_dict.update(rowcount_dict1)

    # print("rowcount_dict ", rowcount_dict)

    columncount_dict = {}
    for selected_column in col_name:

        freq_vals = df[selected_column].value_counts()

        df_Freq = pd.DataFrame(columns=['Values'])
        # df_Freq['Column']=df.columns
        df_Freq['Values'] = freq_vals
        df_Freq = df_Freq.reset_index().rename(columns={'index': 'Columns'})

        # print('df_Freq colnames', df_Freq.columns)

        col_categories = df_Freq['Columns']

        # print('col_categories col_categories===', col_categories)

        df_Freq_json = dict(col_categories)

        # print('df_Freq shapeee', len(df_Freq))

        if selected_column == col_name[0]:
            temp_dict = {len(df_Freq): 'All'}

            df_Freq_json.update(temp_dict)

        # print("df_Freq_json df_Freq_json updatedd colll", df_Freq_json)

        columncount_dict1 = {selected_column: df_Freq_json}
        columncount_dict.update(columncount_dict1)

    # print("columncount_dict ", columncount_dict)

    return rowcount_dict, columncount_dict


def col_class_display():
    filename = "BLS141_Test_Brand.json"
    selected_column = 'Age_Group'

    df = pd.read_json(settings.PYTHONPATH + filename, orient='records', lines=True)

    freq_vals = df[selected_column].value_counts()
    df_Categories = freq_vals.reset_index().rename(columns={'index': 'Categories'})

    # print("Categories",df_Categories)

    # Categories=df_Categories['Categories']
    Categories = df_Categories['Categories'].to_dict()

    Categories_dict = {selected_column: Categories}

    # print('Categories_dict',Categories_dict)

    return Categories_dict


########################## ADDED BY MIHIR PAWAR - 02-05-2023 - INDICES ON COLUMN PERCENT ##############################
##################################### ALLIGN HEADERS ##########################
def allign_headers_condn(cross_df, percent_calc, seperated_flag_row, seperated_flag_col, row_name, col_name):
    if (percent_calc == 'column_percent') or (percent_calc == 'Indices'):
        ######## Allign Columns Total ##############################
        cross_df = allign_headers_fn(cross_df, percent_calc)
        ######## Allign Columns Total ##############################

    elif (percent_calc == 'row_percent'):
        ################### Allign Rows Total ##############################
        cross_df = cross_df.T
        cross_df = allign_headers_fn(cross_df, percent_calc)
        cross_df = cross_df.T
        ################## Allign Rows Total ##############################

    elif (percent_calc == 'table_percent') or (percent_calc == 'actual_count'):

        # cross_df_total_row = cross_df.loc[:, cross_df.columns.get_level_values(1).isin(['Total', 'Total'])]

        # Remove ['total', 'total'] from the index
        # cross_df = cross_df.drop(('Total', 'Total'))
        ####### Allign Columns Total ##############################
        cross_df_cols = allign_headers_fn(cross_df, percent_calc)
        ####### Allign Columns Total ##############################

        ################### Allign Rows Total ##############################
        cross_df_rows = cross_df_cols.T
        cross_df = allign_headers_fn(cross_df_rows, percent_calc)
        cross_df = cross_df.T
        ################### Allign Rows Total ##############################
        # cross_df = pd.concat([cross_df_total_row,cross_df], axis=0)

    if ((len(row_name) > 1) and (len(col_name) > 1)):
        if ((seperated_flag_row == 0 and seperated_flag_col == 0 and percent_calc != 'row_percent') or
                (seperated_flag_row == 1 and seperated_flag_col == 0 and percent_calc != 'row_percent')):
            unique_groups_level0 = cross_df.columns.get_level_values(0).unique().tolist()
            # Remove 'Total' from the original list
            unique_groups_level0.remove('Total')

            # Insert 'Total' at the first position
            unique_groups_level0.insert(0, 'Total')
            print('unique_groups_level0=====', unique_groups_level0)

            cross_df = cross_df.loc[:, unique_groups_level0]

            cross_df = cross_df.reindex(columns=unique_groups_level0, level=0)

        elif ((seperated_flag_row == 0 and seperated_flag_col == 0 and percent_calc == 'row_percent') or
              (seperated_flag_row == 0 and seperated_flag_col == 1 and percent_calc == 'row_percent')):
            print('699===row_perc condn')
            cross_df = cross_df.T
            unique_groups_level0 = cross_df.columns.get_level_values(0).unique().tolist()
            # Remove 'Total' from the original list
            unique_groups_level0.remove('Total')

            # Insert 'Total' at the first position
            unique_groups_level0.insert(0, 'Total')
            print('unique_groups_level0=====', unique_groups_level0)

            cross_df = cross_df.loc[:, unique_groups_level0]

            cross_df = cross_df.reindex(columns=unique_groups_level0, level=0)
            cross_df = cross_df.T

        if ((percent_calc == 'actual_count' or percent_calc == 'table_percent') and seperated_flag_row == 0):
            # cross_df = cross_df.T
            unique_groups_level0_act_tab = cross_df.index.get_level_values(0).unique().tolist()
            print('unique_groups_level0_act_tab', unique_groups_level0_act_tab)

            unique_groups_level0_act_tab.remove('Total')

            # Insert 'Total' at the first position
            unique_groups_level0_act_tab.insert(0, 'Total')

            cross_df = cross_df.loc[unique_groups_level0_act_tab, :]

            cross_df = cross_df.reindex(index=unique_groups_level0_act_tab, level=0)

    return cross_df


def allign_headers_fn(cross_df, percent_calc):
    Columns_df_lst = []
    # Get the level 0 names in the columns
    level_0_names_cols = cross_df.columns.get_level_values(0).unique().tolist()

    cross_df_col_grp = cross_df.groupby(axis=1, level=0)

    # Access the first group
    for cross_df_col_loop in level_0_names_cols:
        cross_df_col = cross_df_col_grp.get_group(cross_df_col_loop)
        cross_df_col_df = pd.DataFrame(cross_df_col)

        unique_groups_level1 = cross_df_col_df.columns.get_level_values(1).unique().tolist()
        level0_grp_names = cross_df_col_df.columns.get_level_values(0).unique().tolist()
        print('level0_grp_names', level0_grp_names)
        print('unique_groups_level1', unique_groups_level1)

        # Remove 'Total' from the original list
        unique_groups_level1.remove('Total')

        # Insert 'Total' at the first position
        unique_groups_level1.insert(0, 'Total')
        print('unique_groups_level1=====', unique_groups_level1)

        cross_df_col_df_reordered = cross_df_col_df.loc[:,
                                    cross_df_col_df.columns.get_level_values(1).isin(list(unique_groups_level1))]
        # cross_df_col_df_reordered.to_excel('cross_df_col_df_reordered_ggg.xlsx')s

        try:
            print('====688 duplicate')
            cross_df_col_df_reordered = cross_df_col_df_reordered.loc[:,
                                        ~cross_df_col_df_reordered.columns.duplicated(keep='first')]
        except:
            pass
        ########################################################################################
        # cross_df_col_df_reordered = cross_df_col_df_reordered.iloc[:, [0, 1]]
        cross_df = cross_df_col_df_reordered.reindex(columns=unique_groups_level1, level=1)
        # cross_df.to_excel('cross_df_col_df_reordered_ffff.xlsx')

        Columns_df_lst.append(cross_df)
    cross_df = pd.concat(Columns_df_lst, axis=1)

    return cross_df


############################### ALLIGN HEADERS BK ########################################################
def remove_duplicate_keys_and_values(dictionary):

    updated_dict = {key: list(set(value)) for key, value in dictionary.items()}
    # updated_dict = {}

    # for key, value in dictionary.items():
    #     value_upd = set(value)
    #     value_upd = list(value_upd)
    #     # print('value_upd',value_upd)
    #     updated_dict[key] = value_upd

    return updated_dict

def remove_prefix(cross_df):
    try:
        cross_df.rename(columns=lambda x: re.sub('.*}', '', x), inplace=True)
    except:
        pass
    try:
        cross_df.rename(index=lambda x: re.sub('.*}', '', x), inplace=True)
    except:
        pass

    # cross_df.rename(columns=lambda x: re.sub('_', ' ', x), inplace=True)
    # cross_df.rename(index=lambda x: re.sub('_', ' ', x), inplace=True)

    # cross_df.rename(columns=lambda x: re.sub('total', '_total', x), inplace=True)
    # cross_df.rename(index=lambda x: re.sub('total', '_total', x), inplace=True)

    return cross_df

####################### ADDED ON 18-02-2025 ################################################
def replace_cy_ya_with_actual_period(cross_df, measure_row_column_position, selected_full_period_str, comparative_full_period_str):
    # Define replacements for 'CY' and 'YA' for both periods
    replacements = {
        'CY': selected_full_period_str,
        '_CY': selected_full_period_str,
        'YA': comparative_full_period_str,
        '_YA': comparative_full_period_str
    }
   
    # Function to apply replacements
    def apply_replacements(axis):
        for key, value in replacements.items():
            if axis == 'columns':
                cross_df.rename(columns=lambda x: re.sub(key, value, x), inplace=True)
            elif axis == 'index':
                cross_df.rename(index=lambda x: re.sub(key, value, x), inplace=True)

 
    # Apply replacements based on position
    if measure_row_column_position == "measure_in_column":
        apply_replacements(axis='columns')
    elif measure_row_column_position == "measure_in_row":
        apply_replacements(axis='index')
   
    return cross_df

# def replace_cy_ya_with_actual_period(cross_df, measure_row_column_position, selected_full_period_str, comparative_full_period_str):
#     # Define replacements for 'CY' and 'YA' for both periods
#     replacements = {
#         'CY': selected_full_period_str,
#         '_CY': selected_full_period_str,
#         'YA': comparative_full_period_str,
#         '_YA': comparative_full_period_str
#     }

#     # Function to apply replacements
#     def apply_replacements(axis):
#         if axis == 'columns':
#             cross_df.rename(columns=lambda x: re.sub('|'.join(re.escape(k) for k in replacements), 
#                                                       lambda m: replacements[m.group(0)], str(x)), inplace=True)
#         elif axis == 'index':
#             cross_df.rename(index=lambda x: re.sub('|'.join(re.escape(k) for k in replacements), 
#                                                    lambda m: replacements[m.group(0)], str(x)), inplace=True)

#     # Apply replacements based on position
#     if measure_row_column_position == "measure_in_column":
#         apply_replacements(axis='columns')
#     elif measure_row_column_position == "measure_in_row":
#         apply_replacements(axis='index')

#     return cross_df

####################### ADDED ON 18-02-2025 ################################################

def replace_cy_ya_with_actual_period_OLD_18022025(cross_df,measure_row_column_position,selected_full_period_str,comparative_full_period_str):
    ################## CODE TO REPLACE CY AND YA WITH ACTUAL PERIOD - 19-04-2024 ####
    if measure_row_column_position == "measure_in_column":
        try:
            cross_df.rename(columns=lambda x: re.sub('CY',selected_full_period_str, x), inplace=True)
            cross_df.rename(columns=lambda x: re.sub('_CY',selected_full_period_str, x), inplace=True)
            cross_df.rename(columns=lambda x: re.sub('YA',comparative_full_period_str, x), inplace=True)
            cross_df.rename(columns=lambda x: re.sub('_YA',comparative_full_period_str, x), inplace=True)
        except:
            pass
    elif measure_row_column_position == "measure_in_row":
        try:
            cross_df.rename(index=lambda x: re.sub('CY',selected_full_period_str, x), inplace=True)
            cross_df.rename(index=lambda x: re.sub('_CY',selected_full_period_str, x), inplace=True)
            cross_df.rename(index=lambda x: re.sub('YA',comparative_full_period_str, x), inplace=True)
            cross_df.rename(index=lambda x: re.sub('_YA',comparative_full_period_str, x), inplace=True)
        except:
            pass

    return cross_df


def replace_sales_currency_with_SALES(selected_weight_column22,cross_df,measure_row_column_position):
    selected_weight_column_str = ','.join([str(elem) for elem in selected_weight_column22])
    print('selected_weight_column_str 881',selected_weight_column_str)
    if measure_row_column_position == "measure_in_column":
        try:
            cross_df.rename(columns=lambda x: re.sub(selected_weight_column_str,'Sales', x), inplace=True)
        except:
            pass
    elif measure_row_column_position == "measure_in_row":
        try:
            cross_df.rename(index=lambda x: re.sub(selected_weight_column_str,'Sales', x), inplace=True)
        except:
            pass

    return cross_df

    ################## CODE TO REPLACE CY AND YA WITH ACTUAL PERIOD - 19-04-2024 ####
def drop_grand_total_from_rows(cross_df):
    try:
        # Identify columns that contain "Grand Total" in any level
        grand_total_index = [col for col in cross_df.index if any('Grand Total' in level for level in col)]

        # Drop the identified columns
        cross_df = cross_df.drop(index=grand_total_index)
    except Exception as e:
        print("An error occurred:", e)

    return cross_df

def drop_grand_total_from_columns(cross_df):
    try:
        # Identify columns that contain "Grand Total" in any level
        grand_total_columns = [col for col in cross_df.columns if any('Grand Total' in level for level in col)]

        # Drop the identified columns
        cross_df = cross_df.drop(columns=grand_total_columns)
    except Exception as e:
        print("An error occurred:", e)

    return cross_df

def code_to_replace_QUARTER(period_to_be_replaced_str,measure_selected_key_val_resp,dict_selected_measures_lst):
    ##### CODE TO REPLACE QUARTER - 19-04-2024 ###########################
    if period_to_be_replaced_str =='QUARTER':
    ################################################################################
        specific_word = period_to_be_replaced_str
        for key, value in measure_selected_key_val_resp.items():
            updated_list = []
            for element in value:
                if specific_word in element:
                    updated_list.append(element.replace(specific_word, ''))
                else:
                    updated_list.append(element)
            measure_selected_key_val_resp[key] = updated_list

        for key, value in dict_selected_measures_lst.items():
            updated_list = []
            for element in value:
                if specific_word in element:
                    updated_list.append(element.replace(specific_word, ''))
                else:
                    updated_list.append(element)
            dict_selected_measures_lst[key] = updated_list

    return measure_selected_key_val_resp,dict_selected_measures_lst

    ################################################################################


def replace_string_in_dict(original_dict, old_string, new_string):
    new_dict = {}
    for key, value in original_dict.items():
        # Check if the old string is present in the key
        if isinstance(key, str) and old_string in key:
            key = key.replace(old_string, new_string)
        # Check if the value is a list
        if isinstance(value, list):
            # Replace old string with new string in each element of the list
            value = [item.replace(old_string, new_string) if isinstance(item, str) else item for item in value]
        elif isinstance(value, str) and old_string in value:
            # If value is a string, replace old string with new string
            value = value.replace(old_string, new_string)
        new_dict[key] = value
    return new_dict
