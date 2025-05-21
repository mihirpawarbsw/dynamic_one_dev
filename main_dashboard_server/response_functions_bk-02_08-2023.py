import json
import os
import re
import time

import pandas as pd
import numpy as np
# from django.shortcuts import render
# from django.http import HttpResponse, JsonResponse
# from django.shortcuts import render, redirect
# from django.conf import settings
# # Create your views here.
# # from CCV_Tool.CCV_Tool.settings import pythonpath
from django.conf import settings

from datetime import date
from datetime import datetime
from io import StringIO

# import win32com.client as win32
# from PIL import Image
# from fpdf import FPDF
from openpyxl.styles import numbers,PatternFill, Border, Side,Font
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from openpyxl import Workbook

# from reportlab.pdfgen import canvas
# from reportlab.lib.pagesizes import letter
from openpyxl.styles import Alignment, Font

# output_pythonpath = r"C:\Users\MihirPawar\Desktop\Python Project BSW\Python Files1\ccv_tool\output_crosstab\\"
# PYTHONPATH = r"C:\Users\MihirPawar\Desktop\Python Project BSW\Python Files1\ccv_tool\ccv_survey_data\electrolux\\"
# PYTHONPATH = r"C:\Users\MihirPawar\Desktop\Python Project BSW\Python Files1\ccv_tool\ccv_survey_data\\"
# MERGED_PYTHONPATH = merged_pythonpath = r"C:\Users\MihirPawar\Desktop\Python Project BSW\Python Files1\ccv_tool\merged_data_files\\"

    ####################### BASE FILTER #####################################################################
def base_filter_data(df,dict_base_filter_data):
    print("==== FUNCTION base_filter_data STARTED!")

    # dict_base_filter = {
    #     'Gender': ['Female'],
    #     # 'Marital_Status': ['Single'],
    #     # 'Age_:_Post_code':['20-29 Years','60-69 Years']
    #     'Age_:_Post_code':['21 – 30 years','36 – 45 years']
    # }
    
    keys = list(dict_base_filter_data.keys())
    values = list(dict_base_filter_data.values())
    print("keys", keys)
    print("values", values)
    
    for loop_keys in range(len(keys)):
        keys_temp = keys[loop_keys]
        values_temp = values[loop_keys]

        print("===keys_temp==",keys_temp)
        print("===values_temp==",values_temp)

        df = df[df[keys_temp].isin(values_temp)]

    df.to_excel('filtered_df.xlsx',index=False)

    # df = df[df['Gender'].isin(['Male'])]
    # df = df[df['GBL Age Groups'].isin(['20-29 Years','60-69 Years'])]
    ####################### BASE FILTER #####################################################################
    print("==== FUNCTION base_filter_data ENDED!")
    return df


def base_filter_resp(df,dict_base_filter,row_name,col_name):
    print('======== base_filter_resp =======')
    print('row_name===',row_name) 
    print('col_name===',col_name) 
    ####################### BASE FILTER #####################################################################
    # dict_base_filter = {

    #     'Electrolux_India_respondent':['Gender','Age_:_Post_code'],
    #     # 'Electrolux_China_respondent': ['Gender']
    # }

    keys = list(dict_base_filter.keys())
    values = list(dict_base_filter.values())[0]

    print('values beforee==',values)
    sum_row_col = row_name + col_name
    print('sum_row_col===',sum_row_col)    
    values = [[x for x in values if x not in sum_row_col]]
    print('values after===',values)

# 
    filter_dict_resp = {}

    for loop_keys in range(len(keys)):
        print('=====loop_keys====',loop_keys)
        table_name = keys[loop_keys]
        table_name = table_name + ".json"
        values_temp = values[loop_keys]
        print('=====values_temp====',values_temp)

        # df = pd.read_json(settings.PYTHONPATH + table_name, orient='records', lines=True)
        df = df[values_temp]

        for selected_column in df.columns:

            df_Freq = pd.DataFrame(columns=['Count'])

            freq_vals = df[selected_column].value_counts()

            df_Freq['Count'] = freq_vals
            df_Freq = df_Freq.reset_index().rename(columns={'index': 'Column_data'})
            df_Freq['Column_data'] = df_Freq['Column_data'].astype(str)
            df_Freq = df_Freq.sort_values('Column_data')
            print("==df_Freq==",df_Freq)

            categories_list = df_Freq['Column_data']

            filter_dict_resp11 = {selected_column:list(categories_list)}
            filter_dict_resp.update(filter_dict_resp11)

    ####################### BASE FILTER #####################################################################

    print('filter_dict_resp====',filter_dict_resp)
    print('======== base_filter_resp =======')
    
    return filter_dict_resp

#################### ADDED ON 25-07-2023 ##########################################

def base_filter_resp_all(df,dict_table,row_name,col_name):
    print('======== base_filter_resp =======')
    # print('row_name===',row_name) 
    # print('col_name===',col_name) 
    ####################### BASE FILTER #####################################################################
    # dict_base_filter = {

    #     'Electrolux_India_respondent':['Gender','Age_:_Post_code'],
    #     # 'Electrolux_China_respondent': ['Gender']
    # }

    keys = list(dict_table.keys())
    values = list(dict_table.values())[0]

    print('keys beforee== aaalll 145',keys)
    print('values beforee== aaalll 146',values)
    sum_row_col = row_name + col_name
    # print('sum_row_col===',sum_row_col)    
    # values = [[x for x in values if x not in sum_row_col]]
    # print('values after===',values)


    filter_dict_resp = {}

    for loop_keys in range(len(keys)):
        print('=====loop_keys====all 155',loop_keys)
        table_name = keys[loop_keys]
        # table_name = table_name + ".json"

        values_temp = dict_table[table_name]

        # values_temp = values[loop_keys]
        # print('=====values_temp====160',values_temp)
        # print('=====values_temp====161',len(values_temp))

        # df = pd.read_json(settings.PYTHONPATH + table_name, orient='records', lines=True)
        df = df[values_temp]

        for selected_column in df.columns:

            df_Freq = pd.DataFrame(columns=['Count'])

            freq_vals = df[selected_column].value_counts()

            df_Freq['Count'] = freq_vals
            df_Freq = df_Freq.reset_index().rename(columns={'index': 'Column_data'})
            df_Freq['Column_data'] = df_Freq['Column_data'].astype(str)
            df_Freq = df_Freq.sort_values('Column_data')
            print("==df_Freq==",df_Freq)

            categories_list = df_Freq['Column_data']

            categories_list = list(categories_list)

            # Insert 'Total' at the beginning of the list
            categories_list.insert(0, 'Total')

            filter_dict_resp11 = {selected_column:categories_list}
            filter_dict_resp.update(filter_dict_resp11)


    ####################### BASE FILTER #####################################################################

    print('filter_dict_resp====',filter_dict_resp)
    print('======== base_filter_resp =======')
    
    return filter_dict_resp
#################### ADDED ON 25-07-2023 ##########################################


def saving_crosstab_excel_file(cross_df, legends_df, percent_calc,indices_calc_flag):
    print('FUNCTION saving_crosstab_excel_file BEGINS...')
    ####################################################### NEW CODE! ##################################
    cross_df_highlighter = cross_df.copy()
    cross_df_highlighter.to_excel(settings.FINAL_CROSSTAB_OUTPUT + 'cross_df_highlighter.xlsx', engine="openpyxl")
    legends_df.to_excel(settings.FINAL_CROSSTAB_OUTPUT + 'legends_df.xlsx', engine="openpyxl")

    workbook_top = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'cross_df_highlighter.xlsx')
    sheet_top = workbook_top.active

    # Open the bottom Excel file
    workbook_bottom = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'legends_df.xlsx')
    sheet_bottom = workbook_bottom.active

    # Apply formatting to the top Excel file based on the calculation type
    if percent_calc == 'column_percent' or percent_calc == 'row_percent' or percent_calc == 'table_percent' or percent_calc == 'Significance':
        percentage_format = numbers.FORMAT_PERCENTAGE_00
        for row in sheet_top.iter_rows():
            for cell in row:
                cell.number_format = percentage_format

    ############# TABLE BORDER ###########################
    for row in sheet_top.iter_rows():
        for cell in row:
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            # Apply the border to the cell
            cell.border = border
    ############## TABLE BORDER ##########################

    ################# FOR INDICES ONLY ###########################################################

    # Get the maximum row index of the top Excel file
    max_row_top = sheet_top.max_row

    # Append the rows from the bottom Excel file to the top Excel file
    for row in sheet_bottom.iter_rows(min_row=2):
        new_row = [cell.value for cell in row]
        sheet_top.append(new_row)

    # Save the combined Excel file
    workbook_top.save(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_CROSSTAB_DF.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################
    workbook = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_CROSSTAB_DF.xlsx')
    # Select the active sheet
    sheet = workbook.active

    # Adjust column width TO 15
    for column in sheet.columns:
        # max_length = 0
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = 15

    # Adjust row height based on cell contents
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                lines = str(cell.value).count('\n') + 1
                cell_height = (lines * 14) + 6
                if cell_height > max_height:
                    max_height = cell_height
            except TypeError:
                pass
        sheet.row_dimensions[row[0].row].height = max_height

    # Left align all cell values and remove bold formatting from text
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top')
            cell.font = Font(bold=False)

    # Save the modified Excel file
    workbook.save(settings.FINAL_CROSSTAB_OUTPUT + 'cross_df.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################
    


################################### SAVING SIGNIFICANCE EXCEL FILE ##################################
def saving_siginficance_excel_file(significance_df):
    print('FUNCTION saving_siginficance_excel_file BEGINS...')
    ####################################################### NEW CODE! ##################################
    significance_df_highlighter = significance_df.copy()
    significance_df_highlighter.to_excel(settings.FINAL_CROSSTAB_OUTPUT + 'significance_df_highlighter.xlsx', engine="openpyxl")

    workbook_top = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    sheet_top = workbook_top.active

    ############# TABLE BORDER ###########################
    for row in sheet_top.iter_rows():
        for cell in row:
            border = Border(left=Side(style='thin'),
                            right=Side(style='thin'),
                            top=Side(style='thin'),
                            bottom=Side(style='thin'))
            # Apply the border to the cell
            cell.border = border
    ############## TABLE BORDER ##########################

    # Get the maximum row index of the top Excel file
    max_row_top = sheet_top.max_row

    # Save the combined Excel file
    workbook_top.save(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################
    workbook = load_workbook(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    # Select the active sheet
    sheet = workbook.active

    # Adjust column width TO 15
    for column in sheet.columns:
        # max_length = 0
        column_letter = get_column_letter(column[0].column)
        sheet.column_dimensions[column_letter].width = 15

    # Adjust row height based on cell contents
    for row in sheet.rows:
        max_height = 0
        for cell in row:
            try:
                lines = str(cell.value).count('\n') + 1
                cell_height = (lines * 14) + 6
                if cell_height > max_height:
                    max_height = cell_height
            except TypeError:
                pass
        sheet.row_dimensions[row[0].row].height = max_height

    # Left align all cell values and remove bold formatting from text
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(horizontal='left', vertical='top')
            cell.font = Font(bold=False)

    # Save the modified Excel file
    workbook.save(settings.FINAL_CROSSTAB_OUTPUT + 'FINAL_SIGNIFICANCE_DF.xlsx')
    ######################## CODE TO MAXIMIZE ALL CELLS, LEFT ALLIGN AND REMOVE BOLD FROM EXCEL ######################

    ################################### SAVING SIGNIFICANCE EXCEL FILE ##################################

def codeframe(list_resp):
    
    # df = pd.read_json(PYTHONPATH + "Electrolux_India.json",orient='records', lines=True)
    # codeframe_df = pd.read_json(PYTHONPATH + "Electrolux- India_codeframe.json",orient='records', lines=True)
    df = pd.DataFrame(list_resp[0])
    codeframe_df = pd.DataFrame(list_resp[1])

    codeframe_df['Variable Level'].fillna('None',inplace=True)
    codeframe_df['Variable Name'].fillna('None',inplace=True)

    encode_dict = {}
    for loop_vals in range(len(codeframe_df)):
        codeframe_dict = {codeframe_df['Variable Name'][loop_vals]:codeframe_df['Variable Level'][loop_vals]}
        encode_dict.update(codeframe_dict)

    df.rename(columns=encode_dict, inplace=True)

    new_col_names_lst = []
    for loop in df.columns:
        # print("loop",loop)
        loop = loop.replace(" ", "_")
        new_col_names_lst.append(loop)

    df.columns = new_col_names_lst
    return df


def column_dtypes(list_all_files):
        colname_dict_final={}
        for loop_filename in list_all_files:

            # pythonpath=r"C:/Users/MihirPawar/Desktop/Python Project BSW/Python Files1/ccv_tool//"
            df = pd.read_json(settings.PYTHONPATH + loop_filename, orient='records', lines=True)
           #print("df", df.head(2))

            # colname_obj = []
            # for x in df.columns:
            #         colname_obj.append(x)

            # colname_obj = [x for x in df.columns]
            # colname_obj = [x for x in df.columns.values]
            colname_obj = list(df.columns.values)
                    
           #print("colname_num", colname_num)
           #print("colname_obj", colname_obj)

            col_dtype_dict = {'filename':loop_filename,'colname_obj': colname_obj}
            filename11=loop_filename.replace(".json","")

            colname_dict_final1={filename11:col_dtype_dict}

           #print("col_dtype_dict",col_dtype_dict)
            colname_dict_final.update(colname_dict_final1)

        return colname_dict_final

def column_data_display(list_all_files):
        colname_data_dict_final = {}
        for loop_filename in list_all_files:
            df=pd.read_json(settings.PYTHONPATH+loop_filename, orient='records',lines=True)
           #print("df", df.head(2))

            colname_data_dict_final_22={}
            # selected_column='Age_Group'
            for selected_column in df.columns:
               #print("selected_column selected_column",selected_column)
                freq_percent=df[selected_column].value_counts(normalize=True).mul(100)
                ##print("freq_percent",freq_percent)

                freq_vals=df[selected_column].value_counts()

                df_Freq=pd.DataFrame(columns=['Count','Percent'])
                # df_Freq['Column']=df.columns
                df_Freq['Count']=freq_vals
                df_Freq['Percent']=freq_percent

                df_Freq['Count']=round(df_Freq['Count'],2)
                df_Freq['Percent']=round(df_Freq['Percent'],2)

                df_Freq['Percent']=df_Freq['Percent'].astype(str) + " %"

                df_Freq=df_Freq.reset_index().rename(columns={'index':'Column_data'})
               #print('df_Freq-->')
               #print('typee df_Freq',type(df_Freq))

                df_Freq_json1 = df_Freq.to_json(orient='index')
                df_Freq_json = json.loads(df_Freq_json1)

                # df_Freq.to_excel('df_Freq.xlsx')

                colname_data_dict_final_selected_column = {selected_column: df_Freq_json}
                colname_data_dict_final_22.update(colname_data_dict_final_selected_column)

            filename22 = loop_filename.replace(".json", "")

            colname_data_dict_final1 = {filename22: colname_data_dict_final_22}
           #print(" ===== resp out of loop ===")
           #print("colname_data_dict_final1", colname_data_dict_final1)

            colname_data_dict_final.update(colname_data_dict_final1)
           #print("======= resp enddd =====")
        return colname_data_dict_final


def value_count_freq(df,filename,row_name,col_name):

    df = pd.read_json(settings.PYTHONPATH + filename, orient='records', lines=True)
    ##print("df", df.head(2))
    ##print("df colsss", df.columns)

    rowcount_dict = {}
    for selected_column in row_name:

        freq_vals = df[selected_column].value_counts()

        df_Freq = pd.DataFrame(columns=['Values'])
        # df_Freq['Column']=df.columns
        df_Freq['Values'] = freq_vals
        df_Freq = df_Freq.reset_index().rename(columns={'index': 'Columns'})

       #print('df_Freq colnames', df_Freq.columns)

        col_categories = df_Freq['Columns']

       #print('col_categories col_categories===', col_categories)

        df_Freq_json = dict(col_categories)

       #print('df_Freq shapeee', len(df_Freq))

        if selected_column == row_name[0]:
            temp_dict = {len(df_Freq): 'All'}

            df_Freq_json.update(temp_dict)

       #print("df_Freq_json df_Freq_json updatedd roww", df_Freq_json)

        rowcount_dict1 = {selected_column: df_Freq_json}
        rowcount_dict.update(rowcount_dict1)

   #print("rowcount_dict ", rowcount_dict)

    columncount_dict = {}
    for selected_column in col_name:

        freq_vals = df[selected_column].value_counts()

        df_Freq = pd.DataFrame(columns=['Values'])
        # df_Freq['Column']=df.columns
        df_Freq['Values'] = freq_vals
        df_Freq = df_Freq.reset_index().rename(columns={'index': 'Columns'})

       #print('df_Freq colnames', df_Freq.columns)

        col_categories = df_Freq['Columns']

       #print('col_categories col_categories===', col_categories)

        df_Freq_json = dict(col_categories)

       #print('df_Freq shapeee', len(df_Freq))

        if selected_column == col_name[0]:
            temp_dict = {len(df_Freq): 'All'}

            df_Freq_json.update(temp_dict)

       #print("df_Freq_json df_Freq_json updatedd colll", df_Freq_json)

        columncount_dict1 = {selected_column: df_Freq_json}
        columncount_dict.update(columncount_dict1)

   #print("columncount_dict ", columncount_dict)

    return rowcount_dict,columncount_dict


def col_class_display():

        filename = "BLS141_Test_Brand.json"
        selected_column = 'Age_Group'

        df=pd.read_json(settings.PYTHONPATH+filename, orient='records',lines=True)

        freq_vals = df[selected_column].value_counts()
        df_Categories = freq_vals.reset_index().rename(columns={'index': 'Categories'})

       #print("Categories",df_Categories)

        # Categories=df_Categories['Categories']
        Categories=df_Categories['Categories'].to_dict()

        Categories_dict={selected_column:Categories}

       #print('Categories_dict',Categories_dict)

        return Categories_dict

########################## ADDED BY MIHIR PAWAR - 02-05-2023 - INDICES ON COLUMN PERCENT ##############################
##################################### ALLIGN HEADERS ##########################
def allign_headers_condn(cross_df,percent_calc,seperated_flag_row,seperated_flag_col,row_name,col_name):
    if (percent_calc == 'column_percent') or (percent_calc == 'Indices'):
        ######## Allign Columns Total ##############################
        cross_df = allign_headers_fn(cross_df,percent_calc)
        ######## Allign Columns Total ##############################

    elif (percent_calc == 'row_percent'):
        ################### Allign Rows Total ##############################
        cross_df = cross_df.T
        cross_df = allign_headers_fn(cross_df,percent_calc)
        cross_df = cross_df.T
        ################## Allign Rows Total ##############################

    elif (percent_calc == 'table_percent') or (percent_calc == 'actual_count'):

        # cross_df_total_row = cross_df.loc[:, cross_df.columns.get_level_values(1).isin(['Total', 'Total'])]

        # Remove ['total', 'total'] from the index
        # cross_df = cross_df.drop(('Total', 'Total'))
        ####### Allign Columns Total ##############################
        cross_df_cols = allign_headers_fn(cross_df,percent_calc)
        ####### Allign Columns Total ##############################

        ################### Allign Rows Total ##############################
        cross_df_rows = cross_df_cols.T
        cross_df = allign_headers_fn(cross_df_rows,percent_calc)
        cross_df = cross_df.T
        ################### Allign Rows Total ##############################
        # cross_df = pd.concat([cross_df_total_row,cross_df], axis=0)


    if ((len(row_name) > 1) and (len(col_name) > 1)):
        if ((seperated_flag_row == 0 and seperated_flag_col == 0 and percent_calc != 'row_percent') or
            (seperated_flag_row == 1 and seperated_flag_col == 0 and percent_calc != 'row_percent')):
            unique_groups_level0 = cross_df.columns.get_level_values(0).unique().tolist()
            # Remove 'Total' from the original list
            unique_groups_level0.remove('Total')

            # Insert 'Total' at the first position
            unique_groups_level0.insert(0, 'Total')
            print('unique_groups_level0=====',unique_groups_level0)

            cross_df = cross_df.loc[:,unique_groups_level0]

            cross_df = cross_df.reindex(columns = unique_groups_level0, level=0)

        elif ((seperated_flag_row == 0 and seperated_flag_col == 0 and percent_calc == 'row_percent') or
            (seperated_flag_row == 0 and seperated_flag_col == 1 and percent_calc == 'row_percent')):
            print('699===row_perc condn')
            cross_df = cross_df.T
            unique_groups_level0 = cross_df.columns.get_level_values(0).unique().tolist()
            # Remove 'Total' from the original list
            unique_groups_level0.remove('Total')

            # Insert 'Total' at the first position
            unique_groups_level0.insert(0, 'Total')
            print('unique_groups_level0=====',unique_groups_level0)

            cross_df = cross_df.loc[:,unique_groups_level0]

            cross_df = cross_df.reindex(columns = unique_groups_level0, level=0)
            cross_df = cross_df.T

        if ((percent_calc == 'actual_count' or percent_calc == 'table_percent') and seperated_flag_row == 0):
            # cross_df = cross_df.T
            unique_groups_level0_act_tab = cross_df.index.get_level_values(0).unique().tolist()
            print('unique_groups_level0_act_tab',unique_groups_level0_act_tab)

            unique_groups_level0_act_tab.remove('Total')

            # Insert 'Total' at the first position
            unique_groups_level0_act_tab.insert(0, 'Total')

            cross_df = cross_df.loc[unique_groups_level0_act_tab,:]

            cross_df = cross_df.reindex(index=unique_groups_level0_act_tab, level=0)

    return cross_df


def allign_headers_fn(cross_df,percent_calc):

    Columns_df_lst = []
    # Get the level 0 names in the columns
    level_0_names_cols = cross_df.columns.get_level_values(0).unique().tolist()

    cross_df_col_grp = cross_df.groupby(axis=1, level=0)

    # Access the first group
    for cross_df_col_loop in level_0_names_cols:
        cross_df_col = cross_df_col_grp.get_group(cross_df_col_loop)
        cross_df_col_df = pd.DataFrame(cross_df_col)

        unique_groups_level1 = cross_df_col_df.columns.get_level_values(1).unique().tolist()
        level0_grp_names = cross_df_col_df.columns.get_level_values(0).unique().tolist()
        print('level0_grp_names', level0_grp_names)
        print('unique_groups_level1', unique_groups_level1)

        # Remove 'Total' from the original list
        unique_groups_level1.remove('Total')

        # Insert 'Total' at the first position
        unique_groups_level1.insert(0, 'Total')
        print('unique_groups_level1=====',unique_groups_level1)

        cross_df_col_df_reordered = cross_df_col_df.loc[:,
                      cross_df_col_df.columns.get_level_values(1).isin(list(unique_groups_level1))]
        cross_df_col_df_reordered.to_excel('cross_df_col_df_reordered_ggg.xlsx')

        try:
            print('====688 duplicate')
            cross_df_col_df_reordered = cross_df_col_df_reordered.loc[:, ~cross_df_col_df_reordered.columns.duplicated(keep='first')]
        except:
            pass
        ########################################################################################
        # cross_df_col_df_reordered = cross_df_col_df_reordered.iloc[:, [0, 1]]
        cross_df = cross_df_col_df_reordered.reindex(columns=unique_groups_level1, level=1)
        cross_df.to_excel('cross_df_col_df_reordered_ffff.xlsx')

        Columns_df_lst.append(cross_df)
    cross_df = pd.concat(Columns_df_lst,axis=1)

    return cross_df
############################### ALLIGN HEADERS BK ########################################################


def remove_prefix(cross_df):
    cross_df.rename(columns=lambda x: re.sub('.*}', '', x), inplace=True)
    cross_df.rename(index=lambda x: re.sub('.*}', '', x), inplace=True)

    cross_df.rename(columns=lambda x: re.sub('_', ' ', x), inplace=True)
    cross_df.rename(index=lambda x: re.sub('_', ' ', x), inplace=True)

    cross_df.rename(columns=lambda x: re.sub('total', '_total', x), inplace=True)
    cross_df.rename(index=lambda x: re.sub('total', '_total', x), inplace=True)

    return cross_df
